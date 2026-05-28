import json
from pathlib import Path
import openpyxl

files = [
    Path(r"C:\Users\lee-y\Downloads\中国制造网回复明细风险筛查模板.xlsx"),
    Path(r"C:\Users\lee-y\Downloads\中国制造平台的原始导出 .xlsx"),
]

for path in files:
    print("\nFILE", path)
    wb = openpyxl.load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        print(json.dumps({"sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column}, ensure_ascii=False))
        for r in range(1, min(ws.max_row, 16) + 1):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 22) + 1)]
            if any(v is not None for v in vals):
                print(r, json.dumps(vals, ensure_ascii=False, default=str))
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((cell.coordinate, cell.value))
        print("FORMULAS", len(formulas))
        for coord, formula in formulas[:80]:
            print(coord, formula)
