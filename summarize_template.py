import json
from pathlib import Path
import openpyxl

path = Path(r"C:\Users\lee-y\Downloads\已经做好的 Excel 分析模板.xlsx")
wb_f = openpyxl.load_workbook(path, data_only=False)
wb_v = openpyxl.load_workbook(path, data_only=True)

for ws in wb_f.worksheets:
    wsv = wb_v[ws.title]
    print("\n##", ws.title, ws.max_row, ws.max_column)
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = []
        for c in range(1, min(ws.max_column, 16) + 1):
            cell = ws.cell(r, c)
            val = wsv.cell(r, c).value
            vals.append(cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else val)
        if any(v is not None for v in vals):
            print(r, json.dumps(vals, ensure_ascii=False, default=str))
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append((cell.coordinate, cell.value))
    print("FORMULA_COUNT", len(formulas))
    for coord, formula in formulas[:60]:
        print(coord, formula)
