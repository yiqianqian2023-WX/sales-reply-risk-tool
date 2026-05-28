from pathlib import Path
from datetime import datetime
import openpyxl

html = Path("index.html").read_text(encoding="utf-8")
for marker in ["多平台回复风险分析工具", "中国制造网", "严重延迟", "回复时长 >= 24 小时"]:
    print("HTML_MARKER", marker, marker in html)

path = Path(r"C:\Users\lee-y\Downloads\中国制造平台的原始导出 .xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
headers = [ws.cell(14, c).value for c in range(1, ws.max_column + 1)]
idx = {name: headers.index(name) + 1 for name in headers if name}

def parse_dt(v):
    if isinstance(v, datetime):
        return v
    return datetime.strptime(str(v), "%Y/%m/%d %H:%M:%S")

records = []
for r in range(15, ws.max_row + 1):
    if not ws.cell(r, idx["发送方首次发送时间"]).value:
        continue
    send = parse_dt(ws.cell(r, idx["发送方首次发送时间"]).value)
    reply = parse_dt(ws.cell(r, idx["首次回复时间"]).value)
    hours = max(0, (reply - send).total_seconds() / 3600)
    fast = 1 if hours <= 4 else 0
    timely_raw = str(ws.cell(r, idx["是否及时回复（24小时）"]).value or "").strip().upper()
    timely = 0 if timely_raw == "N" or hours > 24 else 1
    non_work = 1 if send.isoweekday() > 5 or send.hour < 9 or send.hour >= 18 else 0
    non_work_unfast = 1 if non_work and not fast else 0
    over_one = 1 if hours >= 1 else 0
    over_fast = 1 if hours > 4 else 0
    severe = 1 if hours >= 24 or not timely else 0
    label = ("非工作未极速，" if non_work_unfast else "") + ("超4小时未极速，" if over_fast else "") + ("超1小时，" if over_one else "") + ("超24小时未及时，" if severe else "")
    records.append((ws.cell(r, idx["及时回复负责人"]).value, fast, timely, non_work, non_work_unfast, over_one, over_fast, severe, label))

print("MIC_RECORDS", len(records))
print("MIC_ABNORMAL", sum(1 for row in records if row[-1]))
owners = {}
for owner, fast, timely, non_work, non_work_unfast, over_one, over_fast, severe, label in records:
    owners.setdefault(owner, [0] * 8)
    vals = owners[owner]
    for i, value in enumerate([1, fast, timely, non_work, non_work_unfast, over_one, over_fast, severe]):
        vals[i] += value
print("MIC_OWNERS", len(owners))
