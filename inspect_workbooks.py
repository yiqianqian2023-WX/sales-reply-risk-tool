import json
from pathlib import Path

template_path = Path(r"C:\Users\lee-y\Downloads\已经做好的 Excel 分析模板.xlsx")
raw_path = Path(r"C:\Users\lee-y\Downloads\阿里国际站后台导出的原始回复明细表.xls")

print("PYTHON_IMPORTS")
for name in ["openpyxl", "xlrd", "pandas"]:
    try:
        mod = __import__(name)
        print(name, getattr(mod, "__version__", "ok"))
    except Exception as exc:
        print(name, "ERR", type(exc).__name__, str(exc))

try:
    import openpyxl

    wb_formula = openpyxl.load_workbook(template_path, data_only=False)
    wb_values = openpyxl.load_workbook(template_path, data_only=True)
    print("\nTEMPLATE_SHEETS")
    for ws in wb_formula.worksheets:
        wsv = wb_values[ws.title]
        print(json.dumps({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_ranges": [str(r) for r in list(ws.merged_cells.ranges)[:20]],
        }, ensure_ascii=False))

        rows = []
        for r in range(1, min(ws.max_row, 12) + 1):
            row = []
            for c in range(1, min(ws.max_column, 20) + 1):
                v = wsv.cell(r, c).value
                f = ws.cell(r, c).value
                row.append({"v": v, "f": f if isinstance(f, str) and f.startswith("=") else None})
            rows.append(row)
        print("PREVIEW", json.dumps(rows, ensure_ascii=False, default=str))

        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "value": wsv[cell.coordinate].value,
                    })
        print("FORMULAS", json.dumps(formulas[:120], ensure_ascii=False, default=str))
except Exception as exc:
    print("TEMPLATE_ERR", type(exc).__name__, str(exc))

try:
    import pandas as pd
    xl = pd.ExcelFile(raw_path)
    print("\nRAW_SHEETS", xl.sheet_names)
    for sheet in xl.sheet_names:
        df = pd.read_excel(raw_path, sheet_name=sheet, header=None, nrows=15)
        print("RAW_PREVIEW", sheet, json.dumps(df.where(pd.notnull(df), None).values.tolist(), ensure_ascii=False, default=str))
        df2 = pd.read_excel(raw_path, sheet_name=sheet)
        print("RAW_COLUMNS", sheet, json.dumps(list(df2.columns), ensure_ascii=False, default=str))
        print("RAW_SAMPLE", sheet, json.dumps(df2.head(8).where(pd.notnull(df2), None).to_dict(orient="records"), ensure_ascii=False, default=str))
except Exception as exc:
    print("RAW_ERR", type(exc).__name__, str(exc))
