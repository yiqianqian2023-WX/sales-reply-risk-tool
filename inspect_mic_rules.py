import json
import openpyxl

path = r"C:\Users\lee-y\Downloads\中国制造网回复明细风险筛查模板.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
for sheet_name in ["规则设置", "业务员汇总"]:
    ws = wb[sheet_name]
    print("SHEET", sheet_name)
    for r in range(1, min(ws.max_row, 24) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 18) + 1)]
        print(r, json.dumps(values, ensure_ascii=False, default=str))
